"""
File parsers for the import engine.

Supports CSV (utf-8) and Excel (.xlsx) files.
Both return a list of dicts with string-keyed columns (header row becomes keys).
Empty rows are silently skipped.
"""

import csv
import io
import logging
from typing import IO

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# CSV parser
# ──────────────────────────────────────────────────────────────────────────────

def parse_csv(file: IO) -> list[dict]:
    """
    Parse a CSV file-like object (binary or text).

    • Sniffs delimiter automatically (comma, tab, semicolon).
    • Strips BOM from UTF-8 files saved by Excel.
    • Returns list of dicts; keys are the first-row header values (stripped).
    • Empty rows are omitted.
    """
    if hasattr(file, "read"):
        raw = file.read()
    else:
        raw = file

    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig")   # strip BOM if present

    # Sniff delimiter from the first 4 KB
    sample = raw[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        dialect = csv.excel           # fall back to comma

    reader = csv.DictReader(io.StringIO(raw), dialect=dialect)

    rows = []
    for row in reader:
        # Strip whitespace from keys and values; skip entirely-empty rows
        cleaned = {k.strip(): (v.strip() if v else "") for k, v in row.items() if k}
        if any(cleaned.values()):
            rows.append(cleaned)

    logger.debug("CSV parsed: %d data rows", len(rows))
    return rows


# ──────────────────────────────────────────────────────────────────────────────
# Excel parser
# ──────────────────────────────────────────────────────────────────────────────

def parse_xlsx(file: IO) -> list[dict]:
    """
    Parse an Excel (.xlsx) file-like object.

    • Reads the first worksheet only.
    • Row 1 must be the header row.
    • Cell values are coerced to str (numbers, dates, None).
    • Empty rows (all cells blank) are skipped.
    """
    # Rewind first — the stream may already have been read once this request
    # (e.g. by the column analyser), which would otherwise leave openpyxl an
    # empty buffer and raise "File is not a zip file".
    if hasattr(file, "seek"):
        try:
            file.seek(0)
        except Exception:  # noqa: BLE001
            pass
    raw = file.read() if hasattr(file, "read") else file
    if not raw:
        return []
    if not isinstance(raw, (bytes, bytearray)):
        raw = bytes(raw)

    # Prefer openpyxl; fall back to a styles-free ZIP reader when openpyxl
    # trips over a workbook produced by another tool (e.g. the
    # "expected <class 'openpyxl.styles.fills.Fill'>" error on files written
    # by non-Excel libraries). The fallback ignores styles entirely and reads
    # the raw cell values, so the import keeps working.
    try:
        matrix = _xlsx_rows_via_openpyxl(raw)
    except Exception as exc:  # noqa: BLE001
        logger.warning("openpyxl failed (%s) — using styles-free xlsx reader.", exc)
        matrix = _xlsx_rows_via_zip(raw)

    if not matrix:
        return []

    headers = [str(h).strip() if h not in (None, "") else f"col_{i}"
               for i, h in enumerate(matrix[0])]
    rows = []
    for raw_row in matrix[1:]:
        row = {}
        for key, val in zip(headers, raw_row):
            row[key] = "" if val is None else str(val).strip()
        if any(row.values()):
            rows.append(row)
    logger.debug("XLSX parsed: %d data rows", len(rows))
    return rows


def _xlsx_rows_via_openpyxl(raw: bytes) -> list[list]:
    """Read every row of the active sheet as a 2-D list via openpyxl."""
    import openpyxl  # noqa: PLC0415
    bio = io.BytesIO(raw)
    try:
        wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        bio.seek(0)
        wb = openpyxl.load_workbook(bio, read_only=False, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    matrix = []
    for raw_row in ws.iter_rows(values_only=True):
        cells = []
        for val in raw_row:
            if val is None:
                cells.append("")
            elif isinstance(val, float) and val == int(val):
                cells.append(str(int(val)))    # 1.0 → "1"
            else:
                cells.append(str(val))
        matrix.append(cells)
    wb.close()
    return matrix


def _xlsx_rows_via_zip(raw: bytes) -> list[list]:
    """Pure-stdlib xlsx reader (zipfile + ElementTree) that ignores styles.

    Used when openpyxl rejects an otherwise-valid workbook because of bad
    style definitions written by another tool. Reads the first worksheet's
    cell values, resolving the shared-string table.
    """
    import zipfile  # noqa: PLC0415
    from xml.etree import ElementTree as ET  # noqa: PLC0415

    NS  = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    PR  = "{http://schemas.openxmlformats.org/package/2006/relationships}"
    OR  = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

    zf = zipfile.ZipFile(io.BytesIO(raw))
    names = set(zf.namelist())

    # Shared strings: each <si> may hold one <t> or several <r><t> runs.
    shared = []
    if "xl/sharedStrings.xml" in names:
        sst = ET.fromstring(zf.read("xl/sharedStrings.xml"))
        for si in sst.findall(f"{NS}si"):
            shared.append("".join(t.text or "" for t in si.iter(f"{NS}t")))

    # Resolve the FIRST sheet's part path via workbook rels (fallback sheet1).
    sheet_path = None
    try:
        wbx  = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rid_target = {r.get("Id"): r.get("Target") for r in rels.findall(f"{PR}Relationship")}
        first = list(wbx.find(f"{NS}sheets"))[0]
        target = rid_target.get(first.get(f"{OR}id"))
        if target:
            target = target.lstrip("/")
            sheet_path = target if target.startswith("xl/") else f"xl/{target}"
    except Exception:  # noqa: BLE001
        pass
    if not sheet_path or sheet_path not in names:
        cands = sorted(n for n in names if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"))
        sheet_path = cands[0] if cands else None
    if not sheet_path:
        return []

    def _col_idx(ref: str) -> int:
        letters = "".join(ch for ch in ref if ch.isalpha())
        n = 0
        for ch in letters:
            n = n * 26 + (ord(ch.upper()) - 64)
        return n - 1

    ws = ET.fromstring(zf.read(sheet_path))
    data = ws.find(f"{NS}sheetData")
    if data is None:
        return []

    matrix = []
    for row in data.findall(f"{NS}row"):
        cells = {}
        max_c = -1
        next_c = 0
        for c in row.findall(f"{NS}c"):
            ref = c.get("r") or ""
            idx = _col_idx(ref) if ref else next_c
            next_c = idx + 1
            t = c.get("t")
            val = ""
            if t == "s":
                v = c.find(f"{NS}v")
                if v is not None and v.text is not None:
                    i = int(v.text)
                    val = shared[i] if 0 <= i < len(shared) else ""
            elif t == "inlineStr":
                is_el = c.find(f"{NS}is")
                if is_el is not None:
                    val = "".join(tt.text or "" for tt in is_el.iter(f"{NS}t"))
            else:
                v = c.find(f"{NS}v")
                if v is not None and v.text is not None:
                    val = v.text
                    if val.replace(".", "", 1).lstrip("-").isdigit() and val.endswith(".0"):
                        val = val[:-2]
            cells[idx] = val
            max_c = max(max_c, idx)
        matrix.append([cells.get(i, "") for i in range(max_c + 1)] if max_c >= 0 else [])
    return matrix


# ──────────────────────────────────────────────────────────────────────────────
# Auto-detect parser
# ──────────────────────────────────────────────────────────────────────────────

def parse_file(file: IO, file_name: str) -> list[dict]:
    """
    Select parser based on file extension and return parsed rows.

    Raises ValueError for unsupported formats.
    """
    lower = file_name.lower()
    if lower.endswith(".csv"):
        return parse_csv(file)
    if lower.endswith((".xlsx", ".xls")):
        if lower.endswith(".xls"):
            raise ValueError(
                "Legacy .xls format is not supported. "
                "Please save as .xlsx and re-upload."
            )
        return parse_xlsx(file)
    raise ValueError(
        f"Unsupported file format: '{file_name}'. "
        "Accepted formats: .csv, .xlsx"
    )
