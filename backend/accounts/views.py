"""
API Views — subscribe, payment webhook, set-password, resend link, plans,
           pay-now (renewal), billing status, JWT login.
"""
import logging
from urllib.parse import urlencode
from decimal import Decimal

from drf_spectacular.utils import extend_schema, extend_schema_view, OpenApiParameter, OpenApiTypes
from django.conf import settings
from django.db.models import Q
from django.http import HttpResponseRedirect
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle, ScopedRateThrottle
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenRefreshView as _BaseTokenRefreshView

from . import services
from .coupon_admin import CouponError
from .models import Plan
from .serializers import (
    PlanSerializer,
    SubscribeSerializer,
    SubscribeResponseSerializer,
    PaymentWebhookSerializer,
    SetPasswordSerializer,
    ResendSetupLinkSerializer,
    PayNowSerializer,
    BillingStatusSerializer,
    PaymentHistorySerializer,
    LoginSerializer,
    AdminLoginSerializer,
    TrialSignupSerializer,
)

logger = logging.getLogger(__name__)

# DB is the source of truth for pricing — migration 0008 seeded the canonical
# catalogue. These helpers remain as no-op shims so older imports keep working.
def get_bdt_plan_price(plan) -> Decimal:
    """Return the DB-stored price for a plan (BDT)."""
    return Decimal(str(getattr(plan, "price", 0) or 0))


def _apply_bdt_plan_price(plan: Plan) -> Plan:  # no-op; kept for backwards compat
    return plan


def _mask_phone_simple(phone: str) -> str:
    """Show only the last 3 digits, mask the rest with bullets."""
    if not phone:
        return ""
    digits = "".join(c for c in phone if c.isdigit())
    if len(digits) < 4:
        return "•••"
    return f"{digits[:2]}{'•' * (len(digits) - 5)}{digits[-3:]}"


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/plans/   — public plan list
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Plans"])
class PlanListView(APIView):
    """Return all active subscription plans (no authentication required)."""
    permission_classes = [AllowAny]

    @extend_schema(
        summary="List available plans",
        description="Returns all active subscription plans. Public endpoint — no authentication required.",
        responses={200: PlanSerializer(many=True)},
    )
    def get(self, request):
        plans = list(Plan.objects.filter(is_active=True))
        for p in plans:
            _apply_bdt_plan_price(p)
        return Response(PlanSerializer(plans, many=True).data)


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/auth/check-username/?username=<wanted>
# Live availability check used by the Subscribe form. Public; throttled.
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Auth"])
class UsernameAvailabilityView(APIView):
    """Live username-availability probe used by the Subscribe form."""
    permission_classes = [AllowAny]
    throttle_classes   = [AnonRateThrottle]

    @extend_schema(
        summary="Check whether a username is available",
        description=(
            "Returns `{available: bool, reason: str, suggestions: [str]}`. "
            "When the username is unavailable OR malformed, `suggestions` lists "
            "a few usable alternatives — the frontend can offer them as chips."
        ),
    )
    def get(self, request):
        from . import services as _services  # local — accounts.services already imported elsewhere as `services`

        raw = request.query_params.get("username", "")
        wanted = _services.normalize_username(raw)

        if not wanted:
            return Response({
                "available":   False,
                "reason":      "Enter a username to check.",
                "suggestions": [],
            })

        if not _services.is_valid_username_format(wanted):
            seed = request.query_params.get("seed") or raw or "user"
            return Response({
                "available":   False,
                "reason":      (
                    f"Username must be {_services.USERNAME_MIN_LEN}-"
                    f"{_services.USERNAME_MAX_LEN} characters, start with a "
                    "lowercase letter, and contain only lowercase letters, "
                    "digits, or underscores."
                ),
                "suggestions": _services.suggest_usernames(seed),
            })

        if not _services.is_username_available(wanted):
            # Use the wanted (well-formed) name as the suggestion seed —
            # `wanted_2`, `wanted_3`, `wanted_a4f1` etc.
            return Response({
                "available":   False,
                "reason":      "This username has already been used.",
                "suggestions": _services.suggest_usernames(wanted),
            })

        return Response({
            "available":   True,
            "reason":      "",
            "suggestions": [],
        })


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/subscribe/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Billing"])
class SubscribeView(APIView):
    """
    Step 1 — prospect submits their details and selected plan.
    Returns a gateway URL to redirect the browser to for payment.
    """
    permission_classes = [AllowAny]
    throttle_classes   = [AnonRateThrottle]

    @extend_schema(
        summary="Start a new subscription",
        description=(
            "Create a pending payment record and return the payment gateway URL.\n\n"
            "The browser should be redirected to `payment_url`. "
            "The gateway will call `POST /api/payment/webhook/` with the result."
        ),
        request=SubscribeSerializer,
        responses={201: SubscribeResponseSerializer},
    )
    def post(self, request):
        serializer = SubscribeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        plan = serializer.context["plan"]
        _apply_bdt_plan_price(plan)

        try:
            payment = services.create_pending_payment(
                plan           = plan,
                name           = serializer.validated_data["name"],
                username       = serializer.validated_data["username"],
                email          = serializer.validated_data["email"],
                phone          = serializer.validated_data["phone"],
                business_name  = serializer.validated_data["business_name"],
                address        = serializer.validated_data["address"],
                thana          = serializer.validated_data["thana"],
                district       = serializer.validated_data["district"],
                postal_code    = serializer.validated_data["postal_code"],
                extra_branches = serializer.validated_data.get("extra_branches", 0),
                referral_phone = serializer.validated_data.get("referral_phone", ""),
                coupon_code    = serializer.validated_data.get("coupon_code", ""),
            )
            payment_url = services.build_gateway_url(payment)
        except services.PaymentGatewayError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except CouponError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        response_data = SubscribeResponseSerializer({
            "payment_id":     payment.id,
            "transaction_id": payment.transaction_id,
            "amount":         payment.amount,
            "payment_url":    payment_url,
        }).data

        return Response(response_data, status=status.HTTP_201_CREATED)


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/signup-trial/  — free 14-day trial (no payment)
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Billing"])
class TrialSignupView(APIView):
    """Provision a free 14-day trial account. No payment, no gateway."""
    permission_classes = [AllowAny]
    throttle_classes   = [AnonRateThrottle]

    @extend_schema(
        summary="Start a free 14-day trial",
        description=(
            "Creates the user + tenant + a 14-day trial subscription immediately.\n\n"
            "The user receives a password-setup email and can log in right away. "
            "After 14 days the subscription expires and the account is suspended — "
            "the tenant must then purchase a paid plan to continue."
        ),
        request=TrialSignupSerializer,
        responses={201: None},
    )
    def post(self, request):
        serializer = TrialSignupSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        # Trial signups can omit username — auto-derive from the buyer's
        # display name if blank. Same defaulting for the address fields so
        # the trial form stays short while the model can still hold them.
        from . import services as _svc
        trial_username = serializer.validated_data.get("username") or ""
        if not trial_username:
            seed = serializer.validated_data["name"]
            suggestions = _svc.suggest_usernames(seed, max_suggestions=1)
            trial_username = suggestions[0] if suggestions else _svc.generate_username(seed)

        try:
            result = services.create_trial_account(
                name           = serializer.validated_data["name"],
                username       = trial_username,
                email          = serializer.validated_data["email"],
                phone          = serializer.validated_data["phone"],
                business_name  = serializer.validated_data["business_name"],
                address        = serializer.validated_data.get("address", ""),
                thana          = serializer.validated_data.get("thana", ""),
                district       = serializer.validated_data.get("district", ""),
                postal_code    = serializer.validated_data.get("postal_code", ""),
                referral_phone = serializer.validated_data.get("referral_phone", ""),
            )
        except services.TrialSignupError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        user = result["user"]
        # Optional dev-mode debug payload: surfaces the OTP back to the
        # frontend when SMS_BACKEND=console + DEBUG=True so testing without a
        # SIM works out of the box. In production this is always {}.
        from . import sms as sms_service  # local — avoids cycles
        dev_extra = {}
        if settings.DEBUG and sms_service.backend_name() == "console":
            code = sms_service.get_last_dev_otp(user.id)
            if code:
                dev_extra["_dev_otp"] = code

        return Response(
            {
                "detail":            "Trial activated. We've texted your username and one-time login code to your mobile.",
                "user_id":           str(user.id),
                "username":          user.username,
                "phone_mask":        _mask_phone_simple(user.phone),
                "subscription_id":   str(result["subscription"].id),
                "trial_expires_on":  result["subscription"].next_billing_date.isoformat(),
                "trial_days":        result["subscription"].plan.duration_days,
                **dev_extra,
            },
            status=status.HTTP_201_CREATED,
        )


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/payment/webhook/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Billing"])
class PaymentWebhookView(APIView):
    """
    Step 2 — payment gateway calls this with the payment result.
    """
    permission_classes     = [AllowAny]
    authentication_classes = []
    throttle_classes       = []

    @extend_schema(
        summary="Payment gateway webhook",
        description=(
            "Called by the payment gateway after the user completes (or fails) payment.\n\n"
            "- `SUCCESS` → provisions user account + tenant database (atomic).\n"
            "- `FAILED` → marks payment as failed.\n"
            "- `PENDING` → acknowledged, no action.\n\n"
            "**Security:** HMAC-SHA256 signature is verified via the `X-Signature` header "
            "or the `signature` body field."
        ),
        request=PaymentWebhookSerializer,
        responses={200: None},
        auth=[],
    )
    def post(self, request):
        incoming = request.data if isinstance(request.data, dict) else dict(request.data)
        normalized = services.normalize_webhook_payload(incoming)

        # An SSLCommerz IPN/callback is identified by `tran_id`/`val_id`. These
        # callbacks are public and forgeable, so we DON'T trust the request body
        # — we confirm the transaction server-to-server via SSLCommerz's
        # validator API and use the amount/tran_id it returns.
        is_sslcommerz = ("tran_id" in incoming) or ("val_id" in incoming)

        if is_sslcommerz:
            data = normalized
            # Only the SUCCESS path needs the (paid) validator confirmation;
            # FAILED/CANCELLED callbacks just mark the payment failed.
            if data["payment_status"] == "SUCCESS":
                ok, confirmed = services.validate_sslcommerz_ipn(incoming)
                if not ok:
                    logger.warning("SSLCommerz IPN validation failed for txn=%s", data["transaction_id"])
                    return Response({"detail": "Payment could not be validated."},
                                    status=status.HTTP_401_UNAUTHORIZED)
                # Override with the values SSLCommerz vouched for.
                data["transaction_id"] = confirmed["transaction_id"]
                data["amount"]         = confirmed["amount"]
        else:
            # Generic/legacy webhook shape — strict serializer + HMAC signature.
            serializer = PaymentWebhookSerializer(data=normalized)
            serializer.is_valid(raise_exception=True)
            data = serializer.validated_data

            signature = request.headers.get("X-Signature", "") or data.get("signature", "")
            if not services.verify_webhook_signature(data, signature):
                logger.warning("Invalid webhook signature for txn=%s", data["transaction_id"])
                return Response({"detail": "Invalid signature."}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            ps = data["payment_status"]

            if ps == "SUCCESS":
                result = services.process_webhook_success(
                    transaction_id = data["transaction_id"],
                    amount         = data["amount"],
                )
                return Response(
                    {
                        "detail":          "Payment confirmed. Account provisioned.",
                        "user_id":         str(result["user"].id),
                        "subscription_id": str(result["subscription"].id),
                    },
                    status=status.HTTP_200_OK,
                )

            elif ps == "FAILED":
                services.process_failed_payment(data["transaction_id"])
                return Response({"detail": "Payment marked as failed."}, status=200)

            else:
                return Response({"detail": "Acknowledged — no action for PENDING."}, status=200)

        except services.DuplicateWebhookError as exc:
            logger.info("Duplicate webhook ignored: %s", exc)
            return Response({"detail": str(exc)}, status=status.HTTP_200_OK)

        except services.WebhookError as exc:
            logger.error("Webhook error for txn=%s: %s", data["transaction_id"], exc)
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        except Exception:
            logger.exception("Unexpected webhook crash for txn=%s", data["transaction_id"])
            return Response({"detail": "Internal error."}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/set-password/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Auth"])
class SetPasswordView(APIView):
    """First-login password setup via one-time token."""
    permission_classes = [AllowAny]
    throttle_classes   = [AnonRateThrottle]

    @extend_schema(
        summary="Set initial password",
        description=(
            "Activate a new account by setting the initial password using the "
            "one-time token emailed after subscription payment is confirmed.\n\n"
            "The token is single-use and expires after 24 hours."
        ),
        request=SetPasswordSerializer,
        responses={200: None},
    )
    def post(self, request):
        serializer = SetPasswordSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            user = services.consume_setup_token(
                token_str    = serializer.validated_data["token"],
                new_password = serializer.validated_data["new_password"],
            )
        except services.PasswordSetupError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        from .security_log import record_security_event
        record_security_event("password_set", request=request, actor=user,
                              actor_email=getattr(user, "email", ""))

        return Response(
            {
                "detail":   "Password set successfully. You can now log in.",
                "email":    user.email,
                "username": user.username,
            },
            status=status.HTTP_200_OK,
        )


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/resend-setup-link/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Auth"])
class ResendSetupLinkView(APIView):
    """Re-send the first-login SMS OTP. (Kept at /api/resend-setup-link/ for
    backwards compatibility with the old email-link UI; the new /login-otp
    page POSTs to /api/auth/resend-otp/ instead.)"""
    permission_classes = [AllowAny]
    throttle_classes   = [AnonRateThrottle]

    @extend_schema(
        summary="Resend the first-login SMS OTP (by email)",
        description=(
            "Request a fresh first-login SMS OTP. "
            "Always returns 200 regardless of whether the account exists "
            "(prevents enumeration)."
        ),
        request=ResendSetupLinkSerializer,
        responses={200: None},
    )
    def post(self, request):
        serializer = ResendSetupLinkSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        services.resend_login_otp(email=serializer.validated_data["email"])
        return Response(
            {"detail": "If an account exists for that email, a new login code has been sent by SMS."},
            status=status.HTTP_200_OK,
        )


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/payment-status/<transaction_id>/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Billing"])
class PaymentStatusView(APIView):
    """Poll whether a payment has been confirmed by the gateway."""
    permission_classes = [AllowAny]
    throttle_classes   = []

    @extend_schema(
        summary="Poll payment status",
        description=(
            "Useful when the gateway redirects the browser back before firing the webhook. "
            "The frontend can poll this endpoint until `status` becomes `SUCCESS`."
        ),
        responses={200: None},
    )
    def get(self, request, transaction_id):
        from .models import Payment
        try:
            payment = Payment.objects.get(transaction_id=transaction_id)
        except Payment.DoesNotExist:
            return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

        # Once the payment is SUCCESS the User has been created; surface their
        # username + masked phone so the polling frontend can deep-link straight
        # into the SMS-OTP login screen.
        user = payment.user
        user_block = {}
        if user:
            user_block = {
                "username":   user.username,
                "phone_mask": _mask_phone_simple(user.phone),
            }
            # Dev-mode convenience for testing without a SIM.
            from . import sms as sms_service  # local
            if settings.DEBUG and sms_service.backend_name() == "console":
                code = sms_service.get_last_dev_otp(user.id)
                if code:
                    user_block["_dev_otp"] = code

        # A renewal payment (existing tenant paying their bill) must NOT be
        # treated like a brand-new signup: they already have a password, so the
        # frontend should send them to the normal login page, not the first-
        # login OTP / set-password screen.
        is_renewal = (payment.metadata or {}).get("type") == "renewal"

        return Response({
            "transaction_id": payment.transaction_id,
            "status":         payment.status,
            "amount":         str(payment.amount),
            "paid_at":        payment.paid_at,
            "is_renewal":     is_renewal,
            "user":           user_block,
        })


@extend_schema(tags=["Billing"])
class PaymentReturnView(APIView):
    """
    SSLCommerz success/fail/cancel return URL handler.
    Accepts both GET and POST payloads, then redirects user to frontend status page.
    """
    permission_classes = [AllowAny]
    authentication_classes = []
    throttle_classes = []

    def _redirect(self, request):
        frontend_base = getattr(settings, "FRONTEND_URL", "http://localhost:3050").rstrip("/")

        payload = {}
        payload.update(request.query_params.dict())
        try:
            payload.update(request.data if isinstance(request.data, dict) else dict(request.data))
        except Exception:
            pass

        # Try to finalize payment from return payload itself (SSLCommerz often
        # sends tran_id/status/amount here), so the flow does not depend only
        # on IPN delivery timing.
        normalized = services.normalize_webhook_payload(payload)
        if normalized.get("transaction_id"):
            try:
                ps = str(normalized.get("payment_status", "")).upper()
                if ps == "SUCCESS":
                    # The return URL is a browser redirect — fully forgeable.
                    # Confirm the transaction server-to-server with SSLCommerz
                    # before provisioning, and trust ITS amount/tran_id. If it
                    # can't be validated, do nothing here (the buyer just sees
                    # 'pending'); a genuine payment validates fine.
                    ok, confirmed = services.validate_sslcommerz_ipn(payload)
                    if ok:
                        services.process_webhook_success(
                            transaction_id=confirmed["transaction_id"],
                            amount=confirmed["amount"],
                        )
                    else:
                        logger.warning(
                            "Payment return for txn=%s could not be validated — not provisioning.",
                            normalized["transaction_id"],
                        )
                elif ps == "FAILED":
                    services.process_failed_payment(normalized["transaction_id"])
            except services.DuplicateWebhookError:
                pass
            except Exception:
                logger.exception(
                    "Payment return finalize failed for txn=%s payload=%s",
                    normalized.get("transaction_id"),
                    payload,
                )

        transaction_id = (
            payload.get("txn")
            or payload.get("tran_id")
            or payload.get("transaction_id")
            or ""
        )
        result = (
            str(payload.get("result", "")).lower()
            or str(payload.get("status", "")).lower()
            or "pending"
        )
        query = urlencode({"txn": transaction_id, "result": result})
        return HttpResponseRedirect(f"{frontend_base}/payment/status?{query}")

    @extend_schema(auth=[])
    def get(self, request):
        return self._redirect(request)

    @extend_schema(auth=[])
    def post(self, request):
        return self._redirect(request)


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/auth/login/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Auth"])
class LoginView(APIView):
    """Email + password → JWT access and refresh tokens."""
    permission_classes = [AllowAny]
    throttle_classes   = [ScopedRateThrottle]
    throttle_scope     = "login"

    @extend_schema(
        summary="Login",
        description=(
            "Authenticate with email and password. Returns JWT access and refresh tokens.\n\n"
            "The access token contains: `role`, `permissions`, `email`, `name`, `user_id`.\n\n"
            "Works for both **active** and **suspended** users "
            "(suspended users can still reach billing endpoints).\n\n"
            "Token lifetime: **30 minutes** (access) / **1 day** (refresh)."
        ),
        request=LoginSerializer,
        responses={200: None},
        auth=[],
    )
    def post(self, request):
        from .security_log import record_security_event
        ident = (request.data.get("identifier") or request.data.get("email")
                 or request.data.get("mobile") or request.data.get("phone") or "")
        serializer = LoginSerializer(data=request.data, context={"request": request})
        if not serializer.is_valid():
            record_security_event("login_failure", request=request, success=False,
                                  actor_email=ident)
            from rest_framework.exceptions import ValidationError
            raise ValidationError(serializer.errors)
        record_security_event("login_success", request=request, actor_email=ident)
        return Response(serializer.validated_data, status=status.HTTP_200_OK)


@extend_schema(tags=["Auth"])
class AdminLoginView(APIView):
    """Staff/superuser login endpoint for platform admins."""
    permission_classes = [AllowAny]
    throttle_classes   = [ScopedRateThrottle]
    throttle_scope     = "login"

    @extend_schema(
        summary="Admin login",
        description=(
            "Authenticate platform administrators (staff/superuser) with email and password.\n\n"
            "Non-admin users are denied and should use the tenant login link sent by email."
        ),
        request=AdminLoginSerializer,
        responses={200: None},
        auth=[],
    )
    def post(self, request):
        from .security_log import record_security_event
        ident = (request.data.get("identifier") or request.data.get("email") or "")
        serializer = AdminLoginSerializer(data=request.data, context={"request": request})
        if not serializer.is_valid():
            record_security_event("login_failure", request=request, success=False,
                                  actor_email=ident, detail={"scope": "admin"})
            from rest_framework.exceptions import ValidationError
            raise ValidationError(serializer.errors)
        record_security_event("login_success", request=request, actor_email=ident,
                              detail={"scope": "admin"})
        return Response(serializer.validated_data, status=status.HTTP_200_OK)


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/pay-now/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Billing"])
class PayNowView(APIView):
    """Initiate a renewal payment for an existing subscription."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Initiate renewal payment",
        description=(
            "Create a renewal payment and return the gateway URL.\n\n"
            "Accessible by both **active** and **suspended** users. "
            "The frontend should redirect the browser to `payment_url`."
        ),
        request=PayNowSerializer,
        responses={201: None},
    )
    def post(self, request):
        serializer = PayNowSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            payment = services.create_renewal_payment(user=request.user)
            payment_url = services.build_gateway_url(payment)
        except services.RenewalError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)
        except services.PaymentGatewayError as exc:
            return Response({"detail": str(exc)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {
                "payment_id":     str(payment.id),
                "transaction_id": payment.transaction_id,
                "amount":         str(payment.amount),
                "payment_url":    payment_url,
            },
            status=status.HTTP_201_CREATED,
        )


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/billing/status/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Billing"])
class BillingSummaryView(APIView):
    """Compact billing snapshot — used by the SPA's BillingGate on every nav."""
    permission_classes = [IsAuthenticated]

    @extend_schema(summary="Billing summary", responses={200: None})
    def get(self, request):
        from .serializers import _build_billing_summary
        return Response(_build_billing_summary(request.user))


class BillingStatusView(APIView):
    """Current subscription details and recent payment history."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Get billing status",
        description=(
            "Returns the current subscription details (plan, status, days remaining) "
            "and the 10 most recent payments."
        ),
        responses={200: BillingStatusSerializer},
    )
    def get(self, request):
        from .models import Subscription, Payment

        sub = (
            Subscription.objects
            .filter(user=request.user)
            .select_related("plan")
            .order_by("-created_at")
            .first()
        )
        payments = (
            Payment.objects
            .filter(user=request.user)
            .order_by("-created_at")[:10]
        )
        return Response({
            "subscription": BillingStatusSerializer(sub).data if sub else None,
            "recent_payments": PaymentHistorySerializer(payments, many=True).data,
        })


# ──────────────────────────────────────────────────────────────────────────────
# GET /api/billing/history/
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Billing"])
class PaymentHistoryView(APIView):
    """Full payment history for the authenticated user."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Full payment history",
        description="Returns all payments for the authenticated user, newest first.",
        responses={200: PaymentHistorySerializer(many=True)},
    )
    def get(self, request):
        from .models import Payment
        payments = Payment.objects.filter(user=request.user).order_by("-created_at")
        return Response(PaymentHistorySerializer(payments, many=True).data)


# ──────────────────────────────────────────────────────────────────────────────
# POST /api/auth/token/refresh/
# Thin subclass so drf-spectacular can attach schema metadata to the view.
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Auth"])
@extend_schema(tags=["Auth"])
class LogoutView(APIView):
    """Revoke a refresh token (logout). Idempotent — an already-invalid token
    still returns 200 so the client can clear its session unconditionally."""
    permission_classes     = [AllowAny]   # the refresh token is itself the credential
    authentication_classes = []
    throttle_classes       = []

    @extend_schema(summary="Logout", description="Blacklist the supplied refresh token.", auth=[])
    def post(self, request):
        from rest_framework_simplejwt.tokens import RefreshToken
        from rest_framework_simplejwt.exceptions import TokenError
        token = (request.data or {}).get("refresh", "")
        if not token:
            return Response({"detail": "Refresh token is required."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            RefreshToken(token).blacklist()
        except TokenError:
            pass  # already expired / blacklisted — treat as logged out
        from .security_log import record_security_event
        record_security_event("logout", request=request)
        return Response({"detail": "Logged out."}, status=status.HTTP_200_OK)


class TokenRefreshView(_BaseTokenRefreshView):
    """Refresh JWT access token using a valid refresh token."""

    @extend_schema(
        summary="Refresh access token",
        description=(
            "Exchange a valid refresh token for a new access token.\n\n"
            "The refresh token is rotated on every use — store the new `refresh` value "
            "returned in the response body.\n\n"
            "Token lifetime: **30 minutes** (access) / **1 day** (refresh)."
        ),
        auth=[],
    )
    def post(self, request, *args, **kwargs):
        return super().post(request, *args, **kwargs)


@extend_schema(tags=["Admin"])
class AdminOverviewView(APIView):
    """
    Platform admin overview:
    - all subscribed clients
    - subscription status
    - payment history
    - tenant provisioning state
    - active/suspended users
    """
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Platform admin overview",
        description="Staff/superuser only. Returns cross-tenant client and billing overview from master DB.",
        responses={200: None},
    )
    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        from .models import User, Subscription, Payment, Tenant

        users = (
            User.objects
            .filter(subscriptions__isnull=False)
            .select_related("tenant")
            .distinct()
            .order_by("-created_at")
        )

        clients = []
        for u in users:
            sub = u.subscriptions.select_related("plan").order_by("-created_at").first()
            last_payment = u.payments.order_by("-created_at").first()
            tenant = getattr(u, "tenant", None)
            clients.append({
                "user_id": str(u.id),
                "name": u.name,
                "email": u.email,
                "business_name": u.business_name,
                "user_status": u.status,
                "plan_name": sub.plan.name if sub else None,
                "subscription_status": sub.status if sub else None,
                "next_billing_date": sub.next_billing_date if sub else None,
                "tenant_provisioned": bool(tenant and tenant.is_provisioned),
                "tenant_db_name": tenant.db_name if tenant else None,
                "last_payment_status": last_payment.status if last_payment else None,
                "last_payment_amount": str(last_payment.amount) if last_payment else None,
                "created_at": u.created_at,
            })

        recent_payments = (
            Payment.objects
            .select_related("user", "subscription__plan")
            .order_by("-created_at")[:50]
        )

        response = {
            "summary": {
                "total_clients": users.count(),
                "active_users": User.objects.filter(status="active").count(),
                "suspended_users": User.objects.filter(status="suspended").count(),
                "active_subscriptions": Subscription.objects.filter(status="active").count(),
                "provisioned_tenants": Tenant.objects.filter(is_provisioned=True).count(),
                "pending_tenants": Tenant.objects.filter(is_provisioned=False).count(),
                "total_payments": Payment.objects.count(),
                "successful_payments": Payment.objects.filter(status="success").count(),
                "pending_payments": Payment.objects.filter(status="pending").count(),
                "failed_payments": Payment.objects.filter(status="failed").count(),
            },
            "clients": clients,
            "recent_payments": [
                {
                    "id": str(p.id),
                    "transaction_id": p.transaction_id,
                    "status": p.status,
                    "amount": str(p.amount),
                    "paid_at": p.paid_at,
                    "created_at": p.created_at,
                    "user_email": p.user.email if p.user else None,
                    "metadata_email": (p.metadata or {}).get("email"),
                    # Company + plan so the dashboard's Recent Transactions
                    # feed can show who paid and for which plan.
                    "business_name": (
                        (p.user.business_name if p.user and p.user.business_name else None)
                        or (p.metadata or {}).get("business_name")
                        or (p.user.name if p.user else None)
                        or (p.metadata or {}).get("name")
                    ),
                    "plan_name": (
                        (p.subscription.plan.name if p.subscription and p.subscription.plan else None)
                        or (p.metadata or {}).get("plan_name")
                    ),
                }
                for p in recent_payments
            ],
        }
        return Response(response)


@extend_schema(tags=["Admin"])
class AdminClientsInfoView(APIView):
    """Platform admin "Client's Info" page — every field collected at signup.

    Distinct from AdminOverviewView, which is the dashboard summary +
    recent-payment feed. This one is purpose-built for the directory
    page: full contact info, postal address, and a resolved referrer
    name (looked up by phone snapshot from the Referral row).
    """
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Platform admin — full client directory",
        description=(
            "Staff / superuser only. Returns every signup field — username, "
            "mobile, email, business name, address, thana, district, postal "
            "code, plus referral phone and resolved referrer name when a "
            "Referral row exists for the user."
        ),
        responses={200: None},
    )
    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        from .models import User, Referral

        # Only tenant owners (parent_owner is NULL) — sub-users aren't
        # buyers, they're staff added by a tenant.
        users = (
            User.objects
            .filter(parent_owner__isnull=True, is_superuser=False, is_staff=False)
            .select_related("tenant")
            .prefetch_related("subscriptions__plan")
            .order_by("-created_at")
        )

        # Build a single query for every Referral row keyed by referred_id
        # so we don't N+1 the referrer lookup.
        referrals_by_user = {
            r.referred_id: r
            for r in Referral.objects.select_related("referrer").filter(
                referred_id__in=users.values_list("id", flat=True),
            )
        }

        clients = []
        for u in users:
            sub = u.subscriptions.order_by("-created_at").first()
            ref = referrals_by_user.get(u.id)
            clients.append({
                "user_id":           str(u.id),
                "username":          u.username,
                "name":              u.name,
                "phone":             u.phone,
                "email":             u.email,
                "business_name":     u.business_name,
                "address":           u.address,
                "thana":             u.thana,
                "district":          u.district,
                "postal_code":       u.postal_code,
                "referral_phone":    ref.referrer_phone_snapshot if ref else "",
                "referral_name":     ref.referrer.name if ref and ref.referrer else "",
                "plan_name":         sub.plan.name if sub else None,
                "subscription_status": sub.status if sub else None,
                "user_status":       u.status,
                "created_at":        u.created_at,
            })

        return Response({"clients": clients, "total": len(clients)})


@extend_schema(tags=["Admin"])
class AdminUsersView(APIView):
    """Platform admin: list and create admin/staff users."""
    permission_classes = [IsAuthenticated]

    def _ensure_platform_admin(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        return None

    @extend_schema(
        summary="List platform users",
        description="Returns staff/superuser users from master database.",
        responses={200: None},
    )
    def get(self, request):
        denied = self._ensure_platform_admin(request)
        if denied:
            return denied

        from .models import User

        from .admin_perms import effective_admin_perms

        users = User.objects.filter(Q(is_staff=True) | Q(is_superuser=True)).order_by("-created_at")
        return Response([
            {
                "id": str(u.id),
                "name": u.name,
                "email": u.email,
                "role": u.role,
                "status": u.status,
                "is_staff": u.is_staff,
                "is_superuser": u.is_superuser,
                # Explicitly granted sections (sub-admins) + the effective set
                # (superusers resolve to every section).
                "admin_permissions": list(u.admin_permissions or []),
                "effective_permissions": effective_admin_perms(u),
                "created_at": u.created_at,
            }
            for u in users
        ])

    @extend_schema(
        summary="Create platform admin/staff user",
        description="Create user with role and staff/superuser flags.",
        request=None,
        responses={201: None},
    )
    def post(self, request):
        denied = self._ensure_platform_admin(request)
        if denied:
            return denied

        # Only a superuser may create/grant admin accounts and their
        # section permissions — sub-admins can never escalate access.
        if not request.user.is_superuser:
            return Response({"detail": "Only a super admin can create admin users."},
                            status=status.HTTP_403_FORBIDDEN)

        from .models import User
        from .admin_perms import sanitize_admin_permissions, effective_admin_perms

        name = (request.data.get("name") or "").strip()
        email = (request.data.get("email") or "").strip().lower()
        password = request.data.get("password") or ""
        role = (request.data.get("role") or "admin").strip().lower()
        status_value = (request.data.get("status") or "active").strip().lower()
        is_superuser = bool(request.data.get("is_superuser", False))
        admin_permissions = sanitize_admin_permissions(request.data.get("admin_permissions"))

        if not name or not email or not password:
            return Response(
                {"detail": "name, email, and password are required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if role not in {"owner", "admin", "manager", "cashier"}:
            return Response({"detail": "Invalid role."}, status=status.HTTP_400_BAD_REQUEST)
        if status_value not in {"active", "suspended"}:
            return Response({"detail": "Invalid status."}, status=status.HTTP_400_BAD_REQUEST)
        if User.objects.filter(email__iexact=email).exists():
            return Response({"detail": "Email already exists."}, status=status.HTTP_400_BAD_REQUEST)

        user = User.objects.create_user(
            email=email,
            name=name,
            password=password,
            role=role,
            status=status_value,
            # Every admin-panel user is staff; superuser is the only
            # escalation. Section access comes from admin_permissions.
            is_staff=True,
            is_superuser=is_superuser,
            is_active=(status_value == "active"),
            admin_permissions=([] if is_superuser else admin_permissions),
        )
        return Response(
            {
                "id": str(user.id),
                "email": user.email,
                "name": user.name,
                "role": user.role,
                "status": user.status,
                "is_staff": user.is_staff,
                "is_superuser": user.is_superuser,
                "admin_permissions": list(user.admin_permissions or []),
                "effective_permissions": effective_admin_perms(user),
            },
            status=status.HTTP_201_CREATED,
        )


@extend_schema(tags=["Admin"])
class AdminUserDetailView(APIView):
    """Super admin: edit / remove an existing admin user."""
    permission_classes = [IsAuthenticated]

    def _ensure_superuser(self, request):
        if not request.user.is_superuser:
            return Response({"detail": "Only a super admin can manage admin users."},
                            status=status.HTTP_403_FORBIDDEN)
        return None

    def patch(self, request, pk):
        if (denied := self._ensure_superuser(request)) is not None:
            return denied

        from .models import User
        from .admin_perms import sanitize_admin_permissions, effective_admin_perms

        try:
            user = User.objects.get(id=pk)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)
        if not (user.is_staff or user.is_superuser):
            return Response({"detail": "Not an admin user."}, status=status.HTTP_400_BAD_REQUEST)

        data = request.data or {}
        if "name" in data:
            user.name = (data.get("name") or "").strip() or user.name
        if "role" in data:
            role = (data.get("role") or "").strip().lower()
            if role not in {"owner", "admin", "manager", "cashier"}:
                return Response({"detail": "Invalid role."}, status=status.HTTP_400_BAD_REQUEST)
            user.role = role
        if "status" in data:
            sv = (data.get("status") or "").strip().lower()
            if sv not in {"active", "suspended"}:
                return Response({"detail": "Invalid status."}, status=status.HTTP_400_BAD_REQUEST)
            user.status = sv
            user.is_active = (sv == "active")
        if "is_superuser" in data:
            # Guard: don't let the last superuser demote themselves into
            # lockout. Allow toggling otherwise.
            new_super = bool(data.get("is_superuser"))
            if (not new_super and user.is_superuser
                    and User.objects.filter(is_superuser=True).exclude(id=user.id).count() == 0):
                return Response({"detail": "At least one super admin must remain."},
                                status=status.HTTP_400_BAD_REQUEST)
            user.is_superuser = new_super
            if new_super:
                user.is_staff = True
        if "admin_permissions" in data:
            user.admin_permissions = ([] if user.is_superuser
                                      else sanitize_admin_permissions(data.get("admin_permissions")))
        if data.get("password"):
            if len(data["password"]) < 8:
                return Response({"detail": "Password must be at least 8 characters."},
                                status=status.HTTP_400_BAD_REQUEST)
            user.set_password(data["password"])

        user.save()
        return Response({
            "id": str(user.id),
            "name": user.name,
            "email": user.email,
            "role": user.role,
            "status": user.status,
            "is_staff": user.is_staff,
            "is_superuser": user.is_superuser,
            "admin_permissions": list(user.admin_permissions or []),
            "effective_permissions": effective_admin_perms(user),
        })

    def delete(self, request, pk):
        if (denied := self._ensure_superuser(request)) is not None:
            return denied
        from .models import User
        try:
            user = User.objects.get(id=pk)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)
        if user.id == request.user.id:
            return Response({"detail": "You can't delete your own account."},
                            status=status.HTTP_400_BAD_REQUEST)
        if user.is_superuser and User.objects.filter(is_superuser=True).exclude(id=user.id).count() == 0:
            return Response({"detail": "At least one super admin must remain."},
                            status=status.HTTP_400_BAD_REQUEST)
        user.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@extend_schema(tags=["Admin"])
class AdminPermissionCatalogView(APIView):
    """The list of platform-admin sections that can be granted to a
    sub-admin (drives the checkbox grid on the Admin Users page)."""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)
        from .admin_perms import ADMIN_PERMISSIONS
        return Response(ADMIN_PERMISSIONS)


@extend_schema(tags=["Admin"])
class AdminCancelPaymentView(APIView):
    """Platform admin: delete a pending payment so it stops cluttering the dashboard."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Cancel a pending payment",
        description="Staff/superuser only. Deletes a payment that is still in 'pending' status.",
        responses={200: None, 400: None, 403: None, 404: None},
    )
    def post(self, request, pk):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        from .models import Payment

        try:
            payment = Payment.objects.get(pk=pk)
        except Payment.DoesNotExist:
            return Response({"detail": "Payment not found."}, status=status.HTTP_404_NOT_FOUND)

        if payment.status != Payment.Status.PENDING:
            return Response(
                {"detail": f"Only pending payments can be cancelled (status={payment.status})."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        txn = payment.transaction_id
        payment.delete()
        logger.info("Admin %s cancelled pending payment %s", request.user.email, txn)
        return Response({"detail": f"Pending payment {txn} cancelled."})


@extend_schema(tags=["Admin"])
class AdminProvisionTenantView(APIView):
    """Platform admin: manually trigger tenant DB provisioning for a user."""
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Provision tenant DB for a user",
        description=(
            "Staff/superuser only. Queues the provisioning task for a user whose "
            "tenant DB hasn't been provisioned yet (or failed). Useful as a fallback "
            "when the automatic post-payment provisioning didn't complete."
        ),
        responses={200: None, 400: None, 403: None, 404: None},
    )
    def post(self, request, user_id):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        from .models import User, Tenant
        from .tasks import provision_tenant_db_task

        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        tenant = getattr(user, "tenant", None)
        if tenant and tenant.is_provisioned:
            return Response(
                {"detail": "Tenant is already provisioned."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Try Celery first — if a worker is up, queueing returns instantly
        # and the worker creates the DB + runs migrations in the background.
        # If Celery isn't running (no broker / no worker) the .delay() call
        # either raises OR returns silently with a result we can't track.
        # Either way we fall back to a synchronous provision so the
        # "Provision now" click actually does the work in-process instead
        # of leaving the tenant Pending forever.
        from .tenant_db import provision_tenant  # noqa: PLC0415
        queued = False
        try:
            result = provision_tenant_db_task.delay(str(user.id))
            queued = bool(getattr(result, "id", None))
            if queued:
                logger.info("Admin %s queued provisioning for %s (task id=%s)",
                            request.user.email, user.email, result.id)
        except Exception as exc:
            logger.warning(
                "Celery dispatch failed for admin provision of %s (%s) — "
                "running synchronous fallback.", user.email, exc,
            )

        if not queued:
            try:
                provision_tenant(str(user.id))
            except Exception as exc:
                logger.exception("Synchronous provision failed for %s", user.email)
                return Response(
                    {"detail": f"Provisioning failed: {exc}"},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            user.refresh_from_db()
            tenant = getattr(user, "tenant", None)
            if not tenant or not tenant.is_provisioned:
                return Response(
                    {"detail": "Provision ran but the tenant is still not marked provisioned. Check logs."},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )
            logger.info("Admin %s provisioned %s synchronously",
                        request.user.email, user.email)
            return Response({
                "detail": f"Provisioned {user.email} synchronously.",
                "mode":   "synchronous",
            })

        return Response({
            "detail": f"Provisioning queued for {user.email}.",
            "mode":   "celery",
        })


@extend_schema(tags=["Admin"])
class AdminDeleteClientView(APIView):
    """Platform admin: delete a client (user + cascading subscription, payments, tenant record).

    The tenant database itself is NOT physically dropped — only the metadata
    record is removed. Drop the underlying database manually if needed.
    """
    permission_classes = [IsAuthenticated]

    @extend_schema(
        summary="Delete a client",
        description=(
            "Staff/superuser only. Deletes the user record and cascades to subscription, "
            "payments, and tenant metadata. Does NOT drop the underlying tenant database."
        ),
        responses={200: None, 403: None, 404: None},
    )
    def delete(self, request, user_id):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        from .models import User

        try:
            user = User.objects.get(pk=user_id)
        except User.DoesNotExist:
            return Response({"detail": "User not found."}, status=status.HTTP_404_NOT_FOUND)

        if user.is_staff or user.is_superuser:
            return Response(
                {"detail": "Cannot delete a staff or superuser account."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        email = user.email
        username = user.username
        tenant = getattr(user, "tenant", None)
        tenant_db_name = getattr(tenant, "db_name", None) if tenant else None

        # Free the username + email FIRST so they can be re-used for a brand-new
        # signup, even if a PROTECT-guarded foreign key (audit rows, payments,
        # etc.) blocks the full cascade delete below. Previously a blocked
        # delete left the original row intact, so the name stayed "already
        # used" forever after an admin "deleted" the tenant.
        import uuid as _uuid
        from django.db.models import ProtectedError
        tomb = f"deleted_{_uuid.uuid4().hex[:10]}"
        user.username = tomb
        user.email = f"{tomb}@deleted.invalid"
        user.is_active = False
        try:
            user.save(update_fields=["username", "email", "is_active"])
        except Exception as exc:  # noqa: BLE001
            logger.warning("Could not anonymise client %s before delete: %s", email, exc)

        try:
            user.delete()
        except ProtectedError as exc:
            # Referenced by protected rows — keep the anonymised tombstone so
            # history stays intact while the original username + email are free.
            logger.warning("Hard delete blocked for %s; kept anonymised tombstone: %s", email, exc)

        logger.info(
            "Admin %s deleted client %s / @%s (tenant_db=%s, not physically dropped)",
            request.user.email, email, username, tenant_db_name,
        )
        from .security_log import record_security_event
        record_security_event("tenant_delete", request=request, actor=request.user,
                              actor_email=getattr(request.user, "email", ""),
                              target=email, detail={"username": username, "tenant_db": tenant_db_name})
        return Response({"detail": f"Client {email} deleted. The username '{username}' and email are now free to reuse."})


# ──────────────────────────────────────────────────────────────────────────────
# Platform Notice — admin CRUD + tenant read-only feed
# ──────────────────────────────────────────────────────────────────────────────

from rest_framework import viewsets
from .models import PlatformNotice  # noqa: E402
from .serializers import PlatformNoticeSerializer  # noqa: E402


class PlatformAdminPermission:
    """Reusable permission gate — staff / superuser only."""
    @staticmethod
    def check(request):
        u = getattr(request, "user", None)
        if not u or not u.is_authenticated:
            return False
        return bool(u.is_staff or u.is_superuser)


@extend_schema(tags=["Admin"])
class PlatformNoticeViewSet(viewsets.ModelViewSet):
    """
    Platform-admin CRUD for the notice board.

      GET    /api/admin/notices/        — list (admin sees all, active or not)
      POST   /api/admin/notices/        — create
      GET    /api/admin/notices/<id>/   — retrieve
      PATCH  /api/admin/notices/<id>/   — update
      DELETE /api/admin/notices/<id>/   — delete
    """
    queryset           = PlatformNotice.objects.select_related("created_by").all()
    serializer_class   = PlatformNoticeSerializer
    permission_classes = [IsAuthenticated]

    def check_permissions(self, request):
        super().check_permissions(request)
        if not PlatformAdminPermission.check(request):
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Platform admin only.")

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


def _notice_targets(notice, owner_id: str) -> bool:
    """A notice with an empty target_user_ids is a broadcast (everyone sees
    it). With a non-empty list, only the listed tenant owners — and their
    sub-users, who resolve to the same owner_id — see it."""
    targets = getattr(notice, "target_user_ids", None) or []
    if not targets:
        return True
    return owner_id in {str(t) for t in targets}


@extend_schema(tags=["Notices"])
class TenantActiveNoticesView(APIView):
    """GET /api/notices/active/ — visible notices for the current tenant.

    Returns ONLY rows where is_active=True, published_at<=now, and
    (expires_at is null OR expires_at>now). Sorted newest-first.
    Used by the tenant Dashboard's Notice Board card.

    Any authenticated user (tenant owner, sub-user, even platform admin)
    can read this — it's the same broadcast list for everyone.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.utils import timezone as _tz  # local import — file doesn't import timezone globally
        now = _tz.now()
        qs = (
            PlatformNotice.objects
            .filter(is_active=True, published_at__lte=now)
            .filter(Q(expires_at__isnull=True) | Q(expires_at__gt=now))
            .order_by("-published_at")
        )
        # Targeting: a notice with a non-empty target_user_ids is only for those
        # tenant owners (and their sub-users). The tenant owner is the user's
        # parent_owner (for sub-users) or the user themselves.
        owner_id = str(getattr(request.user, "parent_owner_id", None) or request.user.id)
        visible = [n for n in qs if _notice_targets(n, owner_id)][:10]
        return Response([
            {
                "id":           str(n.id),
                "title":        n.title,
                "body":         n.body,
                "kind":         n.kind,
                "published_at": n.published_at.isoformat(),
                "expires_at":   n.expires_at.isoformat() if n.expires_at else None,
            }
            for n in visible
        ])


# ──────────────────────────────────────────────────────────────────────────────
# Bulk SMS — platform admin broadcast to every tenant
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Bulk SMS"])
class AdminBulkSmsExportView(APIView):
    """GET /api/admin/bulk-sms/export/ — XLSX of every tenant's phone.

    Platform-admin only. Streams an xlsx file ready for download with
    one row per tenant owner: business_name, owner_name, phone, email,
    plan, created. The admin can edit this file (or build a custom
    recipient subset) and re-upload it on the bulk-SMS send page.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        from .models import User
        from django.http import HttpResponse
        from django.utils import timezone as _tz
        import openpyxl

        users = (
            User.objects
            .filter(parent_owner__isnull=True, is_superuser=False, is_staff=False)
            .select_related("tenant")
            .prefetch_related("subscriptions__plan")
            .order_by("business_name", "name")
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clients"
        ws.append([
            "Business name", "Owner name", "Phone", "Email",
            "Username", "Plan", "Created",
        ])
        for u in users:
            sub = u.subscriptions.order_by("-created_at").first()
            ws.append([
                u.business_name or "",
                u.name or "",
                u.phone or "",
                u.email or "",
                u.username or "",
                (sub.plan.name if sub and sub.plan_id else ""),
                u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
            ])
        # Friendly column widths
        for col_letter, width in [("A", 28), ("B", 22), ("C", 18), ("D", 28),
                                  ("E", 18), ("F", 14), ("G", 12)]:
            ws.column_dimensions[col_letter].width = width

        stamp = _tz.now().strftime("%Y%m%d-%H%M")
        filename = f"iffaa-clients-{stamp}.xlsx"
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(resp)
        return resp


@extend_schema(tags=["Bulk SMS"])
class AdminBulkSmsSendView(APIView):
    """POST /api/admin/bulk-sms/send/

    Send a single SMS body to many recipients in one shot. Platform
    admin only. Recipient list comes from (in priority order):

      1. An uploaded xlsx (multipart 'file') — picks the column whose
         header matches 'phone' / 'mobile' / 'msisdn', else the first
         column.
      2. A JSON 'phones': [...] array on the request body.
      3. All tenant owners on the platform (default — sends to every
         active client).

    Body shape (JSON or multipart):
      {
        "message": "Hello — system maintenance tonight at 10 PM.",
        "phones":  ["01711111111", "+8801822222222", ...]  (optional)
      }

    Returns:
      {
        "backend": "ssl_wireless"|"console",
        "sender_id": "IFFAA",
        "total":  N, "sent": K, "failed": N-K,
        "failed_numbers": [...]
      }
    """
    permission_classes = [IsAuthenticated]
    # MultiPartParser so the same endpoint can take an xlsx upload.
    from rest_framework.parsers import MultiPartParser, JSONParser, FormParser
    parser_classes = [JSONParser, FormParser, MultiPartParser]

    def post(self, request):
        if not (request.user.is_staff or request.user.is_superuser):
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

        from .sms import send_sms, backend_name, _normalize_msisdn
        from .models import User
        from django.conf import settings

        message = (request.data.get("message") or "").strip()
        if not message:
            return Response({"detail": "message is required."},
                            status=status.HTTP_400_BAD_REQUEST)
        if len(message) > 1000:
            return Response({"detail": "message is too long (max 1000 chars)."},
                            status=status.HTTP_400_BAD_REQUEST)

        # 1) Uploaded xlsx wins.
        phones: list[str] = []
        uploaded = request.FILES.get("file")
        if uploaded:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(uploaded, read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    header = [str(c or "").strip().lower() for c in rows[0]]
                    phone_idx = None
                    for i, h in enumerate(header):
                        if h in ("phone", "mobile", "msisdn", "number", "phone number"):
                            phone_idx = i
                            break
                    data_rows = rows[1:] if phone_idx is not None else rows
                    if phone_idx is None:
                        phone_idx = 0  # fall back to first column
                    for row in data_rows:
                        if not row: continue
                        if phone_idx >= len(row): continue
                        val = row[phone_idx]
                        if val is None: continue
                        phones.append(str(val).strip())
            except Exception as exc:
                return Response(
                    {"detail": f"Could not read the uploaded xlsx: {exc}"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        # 2) JSON phones array.
        if not phones:
            raw_phones = request.data.get("phones") or []
            if isinstance(raw_phones, list):
                phones = [str(p).strip() for p in raw_phones if str(p).strip()]

        # 3) Default — every tenant owner.
        if not phones:
            phones = list(
                User.objects
                .filter(parent_owner__isnull=True, is_superuser=False,
                        is_staff=False, phone__isnull=False)
                .exclude(phone="")
                .values_list("phone", flat=True)
            )

        # Normalise + dedupe (preserve order). Invalid numbers go
        # straight into failed_numbers without consuming an SMS credit.
        total_input = len(phones)
        seen = set()
        valid: list[str] = []
        invalid_inputs: list[str] = []
        for p in phones:
            n = _normalize_msisdn(p)
            if not n:
                invalid_inputs.append(p)
                continue
            if n in seen:
                continue
            seen.add(n)
            valid.append(n)

        sent = 0
        gateway_failed: list[str] = []
        for n in valid:
            try:
                ok = send_sms(n, message)
            except Exception as exc:
                logger.exception("Bulk SMS failed for %s: %s", n, exc)
                ok = False
            if ok:
                sent += 1
            else:
                gateway_failed.append(n)

        return Response({
            "backend":        backend_name(),
            "sender_id":      getattr(settings, "SSL_WIRELESS_SID", "") or "(console)",
            "total_input":    total_input,
            "attempted":      len(valid),
            "sent":           sent,
            "failed":         len(invalid_inputs) + len(gateway_failed),
            "invalid_inputs": invalid_inputs,
            "failed_numbers": gateway_failed,
        })


@extend_schema(tags=["Notices"])
class MarqueeNoticeView(APIView):
    """GET /api/notices/marquee/ — the currently active marquee notice.

    The platform admin marks one PlatformNotice as is_marquee=True; the
    most recent visible row wins. The frontend renders the result as a
    right-to-left scrolling banner at the top of every tenant page.

    Returns either:
      {"id": ..., "title": ..., "body": ..., "kind": ..., "marquee_speed": N}
    or:
      null   (no active marquee)
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from django.utils import timezone as _tz
        now = _tz.now()
        owner_id = str(getattr(request.user, "parent_owner_id", None) or request.user.id)
        # Return EVERY active marquee notice (newest first), not just one,
        # so the frontend can scroll several announcements back-to-back in
        # a single bar. Previously only the newest was returned, which made
        # a second active marquee make the first one vanish.
        notices = [
            n for n in (
                PlatformNotice.objects
                .filter(is_active=True, is_marquee=True, published_at__lte=now)
                .filter(Q(expires_at__isnull=True) | Q(expires_at__gt=now))
                .order_by("-published_at")
            )
            if _notice_targets(n, owner_id)
        ]
        return Response([
            {
                "id":            str(n.id),
                "title":         n.title,
                "body":          n.body,
                "kind":          n.kind,
                "marquee_speed": n.marquee_speed,
                "published_at":  n.published_at.isoformat(),
                "expires_at":    n.expires_at.isoformat() if n.expires_at else None,
            }
            for n in notices
        ])


# ──────────────────────────────────────────────────────────────────────────────
# Support contact info — env-driven, surfaced to the dashboard
# ──────────────────────────────────────────────────────────────────────────────

@extend_schema(tags=["Notices"])
class SupportInfoView(APIView):
    """GET /api/support/ — support contact info shown on the tenant dashboard.

    Values come from .env so the platform owner can change them without a
    code deploy. Falls back to sensible defaults so dev environments
    don't show empty fields.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response(_support_info_payload())


SUPPORT_KEYS = {
    # config key            → (settings attr,              hard default)
    "support.email":          ("SUPPORT_EMAIL",            "infoiffaa@gmail.com"),
    "support.phone":          ("SUPPORT_PHONE",            "01833387744"),
    "support.office_address": ("SUPPORT_OFFICE_ADDRESS",   ""),
    "support.hours":          ("SUPPORT_HOURS",            "Sun – Thu, 10 am – 6 pm (BD time)"),
}


def _support_info_payload():
    """Resolution order per field:
      1. PlatformConfig row (admin-panel managed — master DB, so one
         update propagates to every tenant instantly).
      2. settings / .env value.
      3. Hard default.
    """
    from django.conf import settings
    from .models import PlatformConfig
    try:
        cfg = {
            c.key: c.value
            for c in PlatformConfig.objects.filter(key__in=SUPPORT_KEYS.keys())
        }
    except Exception:
        cfg = {}
    out = {}
    for key, (attr, default) in SUPPORT_KEYS.items():
        field = key.split(".", 1)[1]
        val = cfg.get(key, "")
        if not val:
            val = getattr(settings, attr, default) or default
        out[field] = val
    return out


class PublicSupportInfoView(APIView):
    """GET /api/public/support/ — support contact info for the public
    website / landing-page chatbot. No authentication required so the
    AI assistant can hand out the support phone + email to prospects.

    Exposes only the public-facing contact fields (email, phone, hours,
    office address) — same source as the authenticated SupportInfoView.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        return Response(_support_info_payload())


class AdminSupportInfoView(APIView):
    """GET/PUT /api/admin/support-info/ — platform admin edits the
    Support card every tenant sees. Values persist in PlatformConfig
    on the master DB."""

    permission_classes = [IsAuthenticated]

    def _guard(self, request):
        return bool(request.user.is_staff or request.user.is_superuser)

    def get(self, request):
        if not self._guard(request):
            return Response({"detail": "Admin only."}, status=status.HTTP_403_FORBIDDEN)
        return Response(_support_info_payload())

    def put(self, request):
        if not self._guard(request):
            return Response({"detail": "Admin only."}, status=status.HTTP_403_FORBIDDEN)
        from .models import PlatformConfig
        data = request.data or {}
        for key in SUPPORT_KEYS:
            field = key.split(".", 1)[1]
            if field in data:
                PlatformConfig.objects.update_or_create(
                    key=key, defaults={"value": str(data[field] or "").strip()},
                )
        return Response(_support_info_payload())
